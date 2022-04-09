import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell

wb = load_workbook('formatting.xlsx')
ws = wb.active
print(ws['A1'].value)

orangeFill = PatternFill(start_color='F58427',
                   end_color='F58427',
                   fill_type='solid')

greenFill = PatternFill(start_color='91D050',
                   end_color='91D050',
                   fill_type='solid')
firstOn = PatternFill(start_color='92CDDC',
                   end_color='92CDDC',
                   fill_type='solid')
secondOn = PatternFill(start_color='D7E2BC',
                   end_color='D7E2BC',
                   fill_type='solid')
postTake = PatternFill(start_color='B1A1C7',
                   end_color='B1A1C7',
                   fill_type='solid')


for i in range(1,102):
	if ws['A{}'.format(i)].value == "Zero":
		ws['A{}'.format(i)].fill = orangeFill
	if ws['A{}'.format(i)].value == "0900-1700":
		ws['A{}'.format(i)].fill = greenFill
	if ws['A{}'.format(i)].value == "0800-1600":
		ws['A{}'.format(i)].fill = postTake
for i in range(1,22):
	if ws['A{}'.format(i)].value == "0900-2130":
		ws['A{}'.format(i)].fill = firstOn
for i in range(23,87):
	if ws['A{}'.format(i)].value == "0900-2130":
		ws['A{}'.format(i)].fill = secondOn
for i in range(87,102):
	if ws['A{}'.format(i)].value == "0900-2130":
		ws['A{}'.format(i)].fill = firstOn



wb.save('formatting.xlsx')